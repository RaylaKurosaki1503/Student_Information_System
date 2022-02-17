###############################################################################
"""
Author: Rayla Kurosaki
File: excel.py
File type: Module
Description: This file contains easy to read functions in order to manipulate
             Workbooks (Microsoft Excel Spreadsheets).
"""
###############################################################################


import openpyxl.utils


def create_workbook():
    """
    Creates a new workbook

    :return: An empty workbook
    """
    return openpyxl.Workbook()


def get_workbook(path):
    """
    Gets the workbook from the path.

    :param path: Absolute path (location) of the file.
    :return: Workbook to manipulate.
    """
    return openpyxl.load_workbook(path)


def save_workbook(wb, filename):
    """
    Saves all changes to the existing workbook if the filename/path is the same
    as the initial filename/path or creates a new workbook if the filename/path
    is different from the initial filename/path.

    :param wb: Current workbook to manipulate.
    :param filename: Name of the saved workbook (filename.xlsx), or the
                     Absolute path to where it should be saved
                     ('C:\...\filename.xlsx').
    :return: None
    """
    wb.save(filename)
    pass


def create_new_worksheet(wb, name):
    """
    Creates a new worksheet and adds it to the end of the workbook.

    :param wb: Current workbook to manipulate.
    :param name: Name of the new worksheet.
    :return: None
    """
    wb.create_sheet(name)
    pass


def get_worksheet_names(wb):
    """
    Gets the names of all the worksheets of the current workbook.

    :param wb: Current workbook to manipulate.
    :return: A list of all the worksheets' names of the current workbook.
    """
    return wb.sheetnames


def get_active_worksheet(wb):
    """
    Gets the active worksheet of the current workbook. The active worksheet is
    the sheet that pops up when you open the workbook.

    :param wb: Current workbook to manipulate.
    :return: The active worksheet of this workbook.
    """
    return wb.active


def get_worksheet(wb, ws_name):
    """
    Returns the worksheet with a specific name.

    :param wb: Current workbook to manipulate.
    :param ws_name: Name of the worksheet to get.
    :return: The worksheet with the correct worksheet name
    """
    return wb[ws_name]


def update_worksheet_name(ws, name):
    """
    Update/Change the name of a specific worksheet.

    :param ws: Current worksheet to manipulate.
    :param name: New name of the worksheet
    :return: None
    """
    ws.title = name
    pass


def get_cell_value(ws, cell):
    """
    Gets the value of a specific cell of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param cell: Cell to get the value of.
    :return: The value of a specific cell of the current worksheet.
    """
    return ws[cell].value


def update_cell_value(ws, cell, value):
    """
    Update/Change the value of a specific cell of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param cell: Cell to change the value of.
    :param value:
    :return: None
    """
    ws[cell] = value
    pass


def add_data_to_worksheet(ws, lst):
    """
    Adds a row of data to the end of the worksheet

    :param ws: Current worksheet to manipulate.
    :param lst: A list of values to add.
    :return: None
    """
    ws.append(lst)
    pass


def insert_empty_row(ws, i):
    """
    Inserts an empty row at the i-th position of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param i: Location of the worksheet to add the empty row (starts at pos 1).
    :return: None
    """
    ws.insert_rows(i)
    pass


def delete_row(ws, i):
    """
    Deletes a row at the i-th position of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param i: Location of the worksheet to delete a row (starts at pos 1).
    :return: None
    """
    ws.delete_rows(i)
    pass


def get_max_rows(ws):
    """
    Gets the number of non-empty rows of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :return: Number of non-empty rows
    """
    return ws.max_row


def insert_empty_col(ws, i):
    """
    Inserts an empty col at the i-th position of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param i: Location of the worksheet to add the empty col (starts at pos 1).
    :return: None
    """
    ws.insert_cols(i)
    pass


def delete_col(ws, i):
    """
    Deletes a column at the i-th position of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :param i: Location of the worksheet to delete a column (starts at pos 1).
    :return: None
    """
    ws.delete_cols(i)
    pass


def get_max_cols(ws):
    """
    Gets the number of non-empty columns of the current worksheet.

    :param ws: Current worksheet to manipulate.
    :return: Number of non-empty columns
    """
    return ws.max_column


def template_iter(ws):
    """
    Template to iterate through each row of the current worksheet
    """
    # Iterate through each row of the worksheet (each row is a list of values).
    for row in ws.values:
        # iterate through each item in a given row
        for value in row:
            continue
    pass
