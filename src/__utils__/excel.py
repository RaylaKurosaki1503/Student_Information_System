"""
Author: Rayla Kurosaki

File: excel.py

Description: This file contains easy to read functions in order to manipulate
             Workbooks (Microsoft Excel Spreadsheets).
"""
import openpyxl.utils
from openpyxl.utils import get_column_letter


def create_workbook():
    """
    Creates a new workbook

    :return: An empty workbook
    """
    return openpyxl.Workbook()


def get_workbook(path):
    """
    Gets the workbook from the path.

    :param path: Absolute path (location) of the file.
    :return: Workbook to manipulate.
    """
    return openpyxl.load_workbook(path)


def save_workbook(workbook, filename):
    """
    Saves all changes to the existing workbook if the filename/path is the
    same as the initial filename/path or creates a new workbook if the
    filename/path is different from the initial filename/path.

    :param workbook: Current workbook to manipulate.
    :param filename: Name of the saved workbook (filename.xlsx), or the
                     Absolute path to where it should be saved
                     ('C:\...\filename.xlsx').
    """
    workbook.save(filename)
    pass


def create_new_worksheet(workbook, worksheet_name, pos=None):
    """
    Creates a new worksheet and either adds it at a specific position, or
    adds it to the end of the workbook if no position is given.

    :param workbook: Current workbook to manipulate.
    :param worksheet_name: Name of the new worksheet.
    :param pos: Inserting the worksheet at position pos.
    """
    # if the user doesn't specify a position to insert the sheet at.
    if pos is None:
        # Insert the sheet at the end of the workbook.
        workbook.create_sheet(worksheet_name)
        pass
    # Otherwise.
    else:
        # Insert the sheet at the user's desired position.
        workbook.create_sheet(worksheet_name, pos)
        pass
    pass


def get_worksheet(workbook, worksheet_name=None):
    """
    Returns the active worksheet or the worksheet with the specified name.

    :param workbook: Current workbook to manipulate.
    :param worksheet_name: Name of the worksheet to get.
    :return: The active worksheet of this workbook or the worksheet with the
             correct worksheet name.
    """
    # If the user doesn't give a worksheet name to search for.
    if worksheet_name is None:
        # Get the active worksheet.
        return workbook.active
    # Otherwise.
    else:
        # Get the worksheet with the correct worksheet name.
        if worksheet_name in get_worksheet_names(workbook):
            return workbook[worksheet_name]
        else:
            return None
    pass


def copy_worksheet(workbook, worksheet):
    """
    Creates a copy of the worksheet.

    :param workbook: The workbook the worksheet is from.
    :param worksheet: Worksheet to copy.
    :return: A copy of the worksheet.
    """
    return workbook.copy_worksheet(worksheet)


def delete_worksheet(workbook, worksheet_name):
    workbook.remove_sheet(get_worksheet(workbook, worksheet_name))
    pass


def get_worksheet_names(workbook):
    """
    Gets the names of all the worksheets of the current workbook.

    :param workbook: Current workbook to manipulate.
    :return: A list of all the worksheets' names of the current workbook.
    """
    return workbook.sheetnames


def update_worksheet_name(worksheet, name):
    """
    Update/Change the name of a specific worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param name: New name of the worksheet
    :return: None
    """
    worksheet.title = name
    pass


def get_cell(worksheet, cell):
    """
    Gets a specific cell of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param cell: A string or a tuple which denotes the cell to access.
    :return: A specific cell of the current worksheet.
    """
    if type(cell) is tuple:
        r, c = cell
        return worksheet.cell(row=r, column=c)
    else:
        return worksheet[cell]
    pass


def get_cell_value(worksheet, cell):
    """
    Gets the value of a specific cell of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param cell: A string or a tuple which denotes the cell to access.
    :return: The value of a specific cell of the current worksheet.
    """
    if type(cell) is tuple:
        r, c = cell
        return worksheet.cell(row=r, column=c).value
    else:
        return worksheet[cell].value
    pass


def update_cell_value(worksheet, cell, value):
    """
    Update/Change the value of a specific cell of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param cell: A string or a tuple which denotes the cell to access.
    :param value: The new value of the cell.
    """
    if type(cell) is tuple:
        r, c = cell
        worksheet.cell(row=r, column=c, value=value)
        pass
    else:
        worksheet[cell] = value
        pass
    pass


def apply_color_scale(worksheet, cell1, cell2, color_rule):
    worksheet.conditional_formatting.add(f"{cell1}:{cell2}", color_rule)
    pass


def merge_cells(worksheet, cell1, cell2):
    """
    Merges the cells from call1 to cell2 in a worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param cell1: First cell in the range.
    :param cell2: Last cell in the range.
    """
    worksheet.merge_cells(f"{cell1}:{cell2}")
    pass


def unmerge_cells(worksheet, cell1, cell2):
    """
    Unmerges the cells from call1 to cell2 in a worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param cell1: First cell in the range.
    :param cell2: Last cell in the range.
    """
    worksheet.unmerge_cells(f"{cell1}:{cell2}")
    pass


def add_data_to_worksheet(worksheet, lst):
    """
    Adds a row of data to the end of the worksheet

    :param worksheet: Current worksheet to manipulate.
    :param lst: A list of values to add.
    """
    worksheet.append(lst)
    pass


def insert_empty_rows(worksheet, pos, n=None):
    """
    Inserts n empty rows at position pos if n is given. If n is not given,
    insert one empty row at position pos instead.

    :param worksheet: Current worksheet to manipulate.
    :param pos: Location of the worksheet to add the empty row (starts at
                pos1).
    :param n: Number of rows to insert from position pos.
    """
    if n is None:
        worksheet.insert_rows(pos)
        pass
    else:
        worksheet.insert_rows(pos, n)
        pass
    pass


def insert_empty_cols(worksheet, pos, n=None):
    """
    Inserts n empty columns at position pos if n is given. If n is not given,
    insert one empty columns at position pos instead.

    :param worksheet: Current worksheet to manipulate.
    :param pos: Location of the worksheet to add the empty columns (starts at
                pos1).
    :param n: Number of columns to insert from position pos.
    """
    if n is None:
        worksheet.insert_cols(pos)
        pass
    else:
        worksheet.insert_cols(pos, n)
        pass
    pass


def delete_row(worksheet, pos, n=None):
    """
    Deletes a row at the i-th position of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param pos: Location of the worksheet to delete a row (starts at pos 1).
    :param n: Number of rows to delete.
    """
    if n is None:
        worksheet.delete_rows(pos)
        pass
    else:
        worksheet.delete_rows(pos, n)
        pass
    pass


def delete_col(worksheet, i):
    """
    Deletes a column at the i-th position of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :param i: Location of the worksheet to delete a column
              (starts at pos 1).
    """
    worksheet.delete_cols(i)
    pass


def get_max_rows(worksheet):
    """
    Gets the number of non-empty rows of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :return: Number of non-empty rows
    """
    return worksheet.max_row


def get_max_cols(worksheet):
    """
    Gets the number of non-empty columns of the current worksheet.

    :param worksheet: Current worksheet to manipulate.
    :return: Number of non-empty columns
    """
    return worksheet.max_column


def autofit_column_width(worksheet, data, col_letters):
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
                    pass
                pass
            else:
                column_widths += [len(cell)]
                pass
            pass
        pass
    for column_width, col_letter in zip(column_widths, col_letters):
        worksheet.column_dimensions[col_letter].width = column_width
        pass
    pass

# def template_iter(ws):
#     """
#     Template to iterate through each row of the current worksheet
#     """
#     # Iterate through each row of the worksheet
#     # (each row is a list of values).
#     for row in ws.values:
#         # iterate through each item in a given row
#         for value in row:
#             continue
#     pass
