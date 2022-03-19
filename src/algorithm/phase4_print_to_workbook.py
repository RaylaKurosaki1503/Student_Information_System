"""
Author: Rayla Kurosaki

File: phase4_print_to_workbook.py

Description:
"""

from openpyxl import styles
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

import copy
from os.path import exists as file_exists

import __utils__ as utils

CYAN = "00FFFF"
PINK = "FFDDF4"


# 10 cols (A:H)


def get_col_width(path):
    """
    Compute the column widths of the soon-to-be-deleted output file.

    :param path: Path to the soon-to-be-deleted output file.
    :return: The column widths of the soon-to-be-deleted output file.
    """
    if not file_exists(path):
        return None
    else:
        n = 10
        col_widths = []
        workbook = utils.get_workbook(path)
        worksheet = utils.get_worksheet(workbook, "Transcript")
        for i in range(1, n + 1):
            col_letter = get_column_letter(i)
            col_widths.append(worksheet.column_dimensions[col_letter].width)
            pass
        return col_widths
    pass


def init_new_workbook(col_widths):
    """
    Initializes a new Microsoft Excel Workbook with the .

    :param col_widths:
    :return:
    """
    workbook = utils.create_workbook()
    worksheet_name = "Transcript"
    utils.create_new_worksheet(workbook, worksheet_name)
    worksheet = utils.get_worksheet(workbook, worksheet_name)
    if col_widths is not None:
        for i, width in enumerate(col_widths):
            col_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[col_letter].width = width
            pass
        pass
    utils.delete_worksheet(workbook, "Sheet")
    return workbook


def add_basic_info(workbook, student):
    font = Font(bold=True, underline="single")
    fill_cyan = styles.PatternFill(
        start_color=CYAN, end_color=CYAN, fill_type="solid"
    )
    fill_pink = styles.PatternFill(
        start_color=PINK, end_color=PINK, fill_type="solid"
    )
    center = Alignment(horizontal="center")

    worksheet = utils.get_worksheet(workbook, "Transcript")

    set_cols_space = ["A", "B", "I", "J"]
    for col in set_cols_space:
        cell_loc = f"{col}1"
        utils.update_cell_value(worksheet, cell_loc, "llll")
        cell = utils.get_cell(worksheet, cell_loc)
        cell.font = Font(color="FFFFFF")

    utils.update_cell_value(worksheet, "C3", "Basic Information")
    utils.merge_cells(worksheet, "C3", "H3")
    cell = utils.get_cell(worksheet, "C3")
    cell.font = font
    cell.fill = fill_cyan
    cell.alignment = center
    utils.update_cell_value(worksheet, "C4", "Name")
    utils.merge_cells(worksheet, "C4", "D4")
    utils.update_cell_value(worksheet, "E4", student.get_name())
    utils.merge_cells(worksheet, "E4", "H4")
    utils.update_cell_value(worksheet, "C5", "Cumulative GPA")
    utils.merge_cells(worksheet, "C5", "D5")
    utils.update_cell_value(worksheet, "E5", student.get_gpa())
    utils.merge_cells(worksheet, "E5", "H5")

    utils.apply_thick_border_style(worksheet, "C3:H5")

    utils.update_cell_value(worksheet, "C6", "Major(s)/Minor(s)")
    utils.merge_cells(worksheet, "C6", "D6")
    cell = utils.get_cell(worksheet, "C6")
    cell.font = font
    cell.fill = fill_cyan

    utils.update_cell_value(worksheet, "E6", "Degree")
    utils.merge_cells(worksheet, "E6", "F6")
    cell = utils.get_cell(worksheet, "E6")
    cell.font = font
    cell.fill = fill_cyan

    utils.update_cell_value(worksheet, "G6", "GPA")
    utils.merge_cells(worksheet, "G6", "H6")
    cell = utils.get_cell(worksheet, "G6")
    cell.font = font
    cell.fill = fill_cyan

    columns = ["C", "E", "G"]
    data = []
    for major in student.majors.values():
        data.append([major.get_name(), major.get_degree(), major.get_gpa()])
        pass
    for minor in student.minors.values():
        data.append([minor.get_name(), "Minor", minor.get_gpa()])
        pass
    for i, data_row in enumerate(data):
        r = i + 7
        for row, col in zip(data_row, columns):
            cell_loc_1 = f"{col}{r}"
            xy = coordinate_from_string(cell_loc_1)
            col = get_column_letter(column_index_from_string(xy[0]) + 1)
            cell_loc_2 = f"{col}{r}"
            utils.update_cell_value(worksheet, cell_loc_1, row)
            utils.merge_cells(worksheet, cell_loc_1, cell_loc_2)
            pass
        pass

    utils.apply_thick_border_style(worksheet, f"C6:H{r}")

    utils.merge_cells(worksheet, "C2", "H2")
    cell = utils.get_cell(worksheet, "C2")
    cell.fill = fill_pink

    utils.merge_cells(worksheet, "B2", f"B{r + 1}")
    cell = utils.get_cell(worksheet, "B2")
    cell.fill = fill_pink

    utils.merge_cells(worksheet, "I2", f"I{r + 1}")
    cell = utils.get_cell(worksheet, "I2")
    cell.fill = fill_pink

    utils.merge_cells(worksheet, f"C{r + 1}", f"H{r + 1}")
    cell = utils.get_cell(worksheet, f"C{r + 1}")
    cell.fill = fill_pink

    utils.apply_thick_border_style(worksheet, f"B2:I{r + 1}")
    pass


def get_course_data(student):
    course_data, gpa_data = {}, {}
    term = ""
    lst = []
    attempted_term, earned_term, units_term, points_term = 0, 0, 0, 0
    attempted_cum, earned_cum, units_cum, points_cum = 0, 0, 0, 0
    for i, course in enumerate(student.get_courses()):
        course_term = course.get_term()
        if course_term != term:
            if i == 0:
                lst.append([
                    course.get_id(), course.get_name(), course.get_credit(),
                    course.get_earned_credit(), course.get_final_grade(),
                    course.get_points()
                ])
                attempted_term += course.get_credit()
                earned_term += course.get_earned_credit()
                units_term += course.get_units()
                points_term += course.get_points()
                attempted_cum += course.get_credit()
                earned_cum += course.get_earned_credit()
                units_cum += course.get_units()
                points_cum += course.get_points()
                pass
            else:
                years, season = term.split(" ")
                course_data[f"{season} Term: {years}"] = lst
                gpa_data[f"{season} Term: {years}"] = [
                    ["Term GPA", points_term / units_term, attempted_term,
                     earned_term, units_term, points_term],
                    ["Cum GPA", points_cum / units_cum, attempted_cum,
                     earned_cum, units_cum, points_cum],
                ]
                lst = [[course.get_id(), course.get_name(),
                        course.get_credit(), course.get_earned_credit(),
                        course.get_final_grade(), course.get_points()]]
                attempted_term = course.get_credit()
                earned_term = course.get_earned_credit()
                units_term = course.get_units()
                points_term = course.get_points()
                attempted_cum += course.get_credit()
                earned_cum += course.get_earned_credit()
                units_cum += course.get_units()
                points_cum += course.get_points()
                pass
            pass
            term = course_term
        else:
            lst.append([
                course.get_id(), course.get_name(), course.get_credit(),
                course.get_earned_credit(), course.get_final_grade(),
                course.get_points()
            ])
            attempted_term += course.get_credit()
            earned_term += course.get_earned_credit()
            units_term += course.get_units()
            points_term += course.get_points()
            attempted_cum += course.get_credit()
            earned_cum += course.get_earned_credit()
            units_cum += course.get_units()
            points_cum += course.get_points()
            pass
        pass
    years, season = term.split(" ")
    course_data[f"{season} Term: {years}"] = lst
    gpa_data[f"{season} Term: {years}"] = [
        ["Term GPA", points_term / units_term, attempted_term, earned_term,
         units_term, points_term],
        ["Cum GPA", points_cum / units_cum, attempted_cum, earned_cum,
         units_cum, points_cum],
    ]
    return course_data, gpa_data


def add_transcript_data(workbook, student):
    font = Font(bold=True, underline="single")
    fill_cyan = styles.PatternFill(
        start_color=CYAN, end_color=CYAN, fill_type="solid"
    )
    fill_pink = styles.PatternFill(
        start_color=PINK, end_color=PINK, fill_type="solid"
    )
    center = Alignment(horizontal="center")

    worksheet = utils.get_worksheet(workbook, "Transcript")
    r = 9 + len(student.get_majors()) + len(student.get_minors())
    course_data, gpa_data = get_course_data(student)
    header = ["Course", "Description", "Attempted", "Earned", "Grade",
              "Points"]
    old_r = copy.deepcopy(r)
    for term, data_set in course_data.items():
        term_gpa, cum_gpa = gpa_data[term]
        cell_loc = f"C{r}"
        utils.update_cell_value(worksheet, cell_loc, term)
        cell = utils.get_cell(worksheet, cell_loc)
        cell.alignment = center
        r += 1
        for j in range(3, 8 + 1):
            cell_loc = f"{get_column_letter(j)}{r}"
            utils.update_cell_value(worksheet, cell_loc, header[j - 3])
            cell = utils.get_cell(worksheet, cell_loc)
            cell.font = font
            cell.fill = fill_cyan
            pass
        r += 1
        for course_data in data_set:
            for j in range(3, 8 + 1):
                c = get_column_letter(j)
                val = course_data[j - 3]
                utils.update_cell_value(worksheet, f"{c}{r}", val)
                pass
            r += 1
            pass
        cell_loc = f"C{old_r + 1}:H{r - 1}"
        utils.apply_thick_border_style(worksheet, cell_loc)
        for j in range(3, 8 + 1):
            cell_loc = f"{get_column_letter(j)}{r}"
            utils.update_cell_value(worksheet, cell_loc, term_gpa[j - 3])
            cell = utils.get_cell(worksheet, cell_loc)
            cell.fill = fill_cyan
            if (j - 3 == 1) or (j - 3 == 5):
                cell.number_format = '0.000'
                pass
            pass
        cell_loc = f"C{r}:H{r}"
        utils.apply_thick_border_style(worksheet, cell_loc)
        r += 1
        for j in range(3, 8 + 1):
            cell_loc = f"{get_column_letter(j)}{r}"
            utils.update_cell_value(worksheet, cell_loc, cum_gpa[j - 3])
            cell = utils.get_cell(worksheet, cell_loc)
            cell.fill = fill_cyan
            if (j - 3 == 1) or (j - 3 == 5):
                cell.number_format = '0.000'
                pass
            pass
        cell_loc = f"C{r}:H{r}"
        utils.apply_thick_border_style(worksheet, cell_loc)
        r += 1
        utils.merge_cells(worksheet, f"B{old_r}", f"B{r}")
        cell = utils.get_cell(worksheet, f"B{old_r}")
        cell.fill = fill_pink
        utils.merge_cells(worksheet, f"I{old_r}", f"I{r}")
        cell = utils.get_cell(worksheet, f"I{old_r}")
        cell.fill = fill_pink
        utils.merge_cells(worksheet, f"C{old_r}", f"H{old_r}")
        cell = utils.get_cell(worksheet, f"C{old_r}")
        cell.fill = fill_pink
        utils.merge_cells(worksheet, f"C{r}", f"H{r}")
        cell = utils.get_cell(worksheet, f"C{r}")
        cell.fill = fill_pink
        cell_range = f"B{old_r}:I{r}"
        utils.apply_thick_border_style(worksheet, cell_range)
        r += 2
        old_r = copy.deepcopy(r)
        pass
    pass


def phase4_main(student):
    path = "../data/transcript.xlsx"
    col_widths = get_col_width(path)
    workbook = init_new_workbook(col_widths)
    add_basic_info(workbook, student)
    add_transcript_data(workbook, student)
    utils.save_workbook(workbook, path)
    pass
